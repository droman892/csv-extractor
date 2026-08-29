from pathlib import Path
import shutil
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


class ResultsExportService:

    HEADER_FILL = "D9EAF7"
    ALTERNATE_FILL = "F2F2F2"
    WHITE_FILL = "FFFFFF"
    BORDER_COLOR = "D0D0D0"

    @staticmethod
    def create_export_file(result):
        temp_directory = Path(
            tempfile.mkdtemp(
                prefix="csv_extractor_"
            )
        )

        export_path = (
            temp_directory
            / f"{Path(result['filename']).stem}_results.xlsx"
        )

        ResultsExportService.export_results(
            result,
            export_path
        )

        return str(export_path)

    @staticmethod
    def export_results(result, destination_path):
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        customer_sheet = workbook.create_sheet(
            "Hours by Customer"
        )

        validation_sheet = workbook.create_sheet(
            "Validation Issues"
        )

        ResultsExportService._build_summary_sheet(
            summary_sheet,
            result
        )

        ResultsExportService._build_customer_sheet(
            customer_sheet,
            result
        )

        ResultsExportService._build_validation_sheet(
            validation_sheet,
            result
        )

        workbook.save(destination_path)

    @staticmethod
    def _apply_row_style(
        sheet,
        row_number,
        start_column,
        end_column,
        data_row_index
    ):
        fill = PatternFill(
            fill_type="solid",
            fgColor=(
                ResultsExportService.ALTERNATE_FILL
                if data_row_index % 2 == 0
                else ResultsExportService.WHITE_FILL
            )
        )

        alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for column in range(
            start_column,
            end_column + 1
        ):
            cell = sheet.cell(
                row=row_number,
                column=column
            )

            cell.fill = fill
            cell.alignment = alignment

    @staticmethod
    def _apply_header_style(
        sheet,
        row_number,
        start_column,
        end_column
    ):
        fill = PatternFill(
            fill_type="solid",
            fgColor=ResultsExportService.HEADER_FILL
        )

        font = Font(
            bold=True
        )

        alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        border = Border(
            bottom=Side(
                style="thin",
                color=ResultsExportService.BORDER_COLOR
            )
        )

        for column in range(
            start_column,
            end_column + 1
        ):
            cell = sheet.cell(
                row=row_number,
                column=column
            )

            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border

    @staticmethod
    def _build_summary_sheet(sheet, result):
        left_alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        section_fill = PatternFill(
            fill_type="solid",
            fgColor=ResultsExportService.HEADER_FILL
        )

        section_font = Font(
            bold=True
        )

        border = Border(
            bottom=Side(
                style="thin",
                color=ResultsExportService.BORDER_COLOR
            )
        )

        # Overall
        sheet["A1"] = "Overall"

        for column in range(1, 3):
            cell = sheet.cell(
                row=1,
                column=column
            )

            cell.fill = section_fill
            cell.border = border

        sheet["A1"].font = section_font
        sheet["A1"].alignment = left_alignment

        sheet["A2"] = "Metric"
        sheet["B2"] = "Count"

        ResultsExportService._apply_header_style(
            sheet,
            2,
            1,
            2
        )

        overall_rows = [
            (
                "Total Tickets",
                result["total_tickets_count"]
            ),
            (
                "Valid Tickets",
                result["valid_tickets_count"]
            ),
            (
                "Invalid Tickets",
                result["invalid_tickets_count"]
            ),
            (
                "Total Hours",
                result["summary"]["total_hours"]
            ),
        ]

        row = 3

        for data_row_index, (metric, value) in enumerate(
            overall_rows
        ):
            sheet.cell(
                row=row,
                column=1,
                value=metric
            )

            sheet.cell(
                row=row,
                column=2,
                value=value
            )

            ResultsExportService._apply_row_style(
                sheet,
                row,
                1,
                2,
                data_row_index
            )

            sheet.cell(
                row=row,
                column=1
            ).alignment = left_alignment

            if metric == "Total Hours":
                sheet.cell(
                    row=row,
                    column=2
                ).number_format = '#,##0.0'
            else:
                sheet.cell(
                    row=row,
                    column=2
                ).number_format = '#,##0'

            row += 1

        row += 1

        # Tickets by Status
        sheet.cell(
            row=row,
            column=1,
            value="Tickets by Status"
        )

        for column in range(1, 3):
            cell = sheet.cell(
                row=row,
                column=column
            )

            cell.fill = section_fill
            cell.border = border

        sheet.cell(
            row=row,
            column=1
        ).font = section_font

        sheet.cell(
            row=row,
            column=1
        ).alignment = left_alignment

        row += 1

        sheet.cell(
            row=row,
            column=1,
            value="Status"
        )

        sheet.cell(
            row=row,
            column=2,
            value="Count"
        )

        ResultsExportService._apply_header_style(
            sheet,
            row,
            1,
            2
        )

        row += 1

        for data_row_index, (status, count) in enumerate(
            result["summary"]["tickets_by_status"].items()
        ):
            sheet.cell(
                row=row,
                column=1,
                value=status.replace(
                    "_",
                    " "
                ).title()
            )

            sheet.cell(
                row=row,
                column=2,
                value=count
            )

            ResultsExportService._apply_row_style(
                sheet,
                row,
                1,
                2,
                data_row_index
            )

            sheet.cell(
                row=row,
                column=2
            ).number_format = '#,##0'

            row += 1

        row += 1

        # Tickets by Priority
        sheet.cell(
            row=row,
            column=1,
            value="Tickets by Priority"
        )

        for column in range(1, 3):
            cell = sheet.cell(
                row=row,
                column=column
            )

            cell.fill = section_fill
            cell.border = border

        sheet.cell(
            row=row,
            column=1
        ).font = section_font

        sheet.cell(
            row=row,
            column=1
        ).alignment = left_alignment

        row += 1

        sheet.cell(
            row=row,
            column=1,
            value="Priority"
        )

        sheet.cell(
            row=row,
            column=2,
            value="Count"
        )

        ResultsExportService._apply_header_style(
            sheet,
            row,
            1,
            2
        )

        row += 1

        for data_row_index, (priority, count) in enumerate(
            result["summary"]["tickets_by_priority"].items()
        ):
            sheet.cell(
                row=row,
                column=1,
                value=priority.replace(
                    "_",
                    " "
                ).title()
            )

            sheet.cell(
                row=row,
                column=2,
                value=count
            )

            ResultsExportService._apply_row_style(
                sheet,
                row,
                1,
                2,
                data_row_index
            )

            sheet.cell(
                row=row,
                column=2
            ).number_format = '#,##0'

            row += 1

        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 18

        sheet.freeze_panes = None

    @staticmethod
    def _build_customer_sheet(sheet, result):
        sheet.append([
            "Customer",
            "Hours"
        ])

        ResultsExportService._apply_header_style(
            sheet,
            1,
            1,
            2
        )

        customers = sorted(
            result["summary"]["hours_by_customer"].items(),
            key=lambda item: item[0].lower()
        )

        for data_row_index, (customer, hours) in enumerate(
            customers
        ):
            sheet.append([
                customer,
                hours
            ])

            row = sheet.max_row

            ResultsExportService._apply_row_style(
                sheet,
                row,
                1,
                2,
                data_row_index
            )

            sheet.cell(
                row=row,
                column=2
            ).number_format = '#,##0.0'

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 15

        if sheet.max_row >= 1:
            sheet.auto_filter.ref = (
                f"A1:B{sheet.max_row}"
            )

    @staticmethod
    def _build_validation_sheet(sheet, result):
        headers = [
            "Issue #",
            "Ticket",
            "Field",
            "Invalid Value",
            "Validation Error"
        ]

        sheet.append(headers)

        ResultsExportService._apply_header_style(
            sheet,
            1,
            1,
            5
        )

        issue_number = 1
        data_row_index = 0

        for record in result["invalid_records"]:
            for error in record["errors"]:
                sheet.append([
                    issue_number,
                    record["ticket_id"],
                    error["field"],
                    error["invalid_value"],
                    error["reason"]
                ])

                row = sheet.max_row

                ResultsExportService._apply_row_style(
                    sheet,
                    row,
                    1,
                    5,
                    data_row_index
                )

                issue_number += 1
                data_row_index += 1

        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 25
        sheet.column_dimensions["E"].width = 50

        if sheet.max_row >= 1:
            sheet.auto_filter.ref = (
                f"A1:E{sheet.max_row}"
            )

    @staticmethod
    def copy_export_file(source_path, destination_path):
        shutil.copyfile(
            source_path,
            destination_path
        )